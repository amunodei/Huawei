"""
CSR Initial Script Generator - Flask web app.
"""
from __future__ import annotations

import io
import os
import re
import zipfile
from datetime import datetime
from typing import Any

import openpyxl
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

CIDR_MASK = {
    "/30": "255.255.255.252",
    "/31": "255.255.255.254",
    "/32": "255.255.255.255",
    "/29": "255.255.255.248",
    "/28": "255.255.255.240",
    "/27": "255.255.255.224",
    "/26": "255.255.255.192",
    "/25": "255.255.255.128",
    "/24": "255.255.255.0",
}

_HERE = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_TEMPLATE_PATH = os.path.join(_HERE, "default_template.txt")


def _load_default_template() -> str:
    try:
        with open(_DEFAULT_TEMPLATE_PATH, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "sys\nsysname {{csr_id}}\n# default_template.txt missing\n"


DEFAULTS = {
    "csr_sheet": "National CSR List",
    "ptp_sheet": "National PTP",
    "header_row": 2,
    "csr_id_col": "CSR_ID",
    "esn_col": "ESN",
    "lo0_ip_col": "Lo0_IP",
    "lo0_cidr_col": "Lo0_CIDR",
    "lo1_ip_col": "Lo1_IP",
    "lo1_cidr_col": "Lo1_CIDR",
    "ptp_csr_id_col": "CSR_ID",
    "ptp_ip_col": "CSR_IP",
    "ptp_cidr_col": "CSR_CIDR",
}


def derive_net_from_ip(ip, area="49.0001", nsel="00"):
    if not ip:
        return ""
    s = str(ip).strip()
    parts = s.split(".")
    if len(parts) != 4:
        return ""
    try:
        octets = [int(p) for p in parts]
    except ValueError:
        return ""
    if not all(0 <= o <= 255 for o in octets):
        return ""
    twelve = "".join("{:03d}".format(o) for o in octets)
    return "{}.{}.{}.{}.{}".format(area, twelve[0:4], twelve[4:8], twelve[8:12], nsel)


def cidr_to_mask(cidr):
    if cidr is None:
        return ""
    s = str(cidr).strip()
    if s in CIDR_MASK:
        return CIDR_MASK[s]
    if "." in s:
        return s
    m = re.match(r"/?(\d{1,2})$", s)
    if m:
        n = int(m.group(1))
        if 0 <= n <= 32:
            mask_int = (0xFFFFFFFF << (32 - n)) & 0xFFFFFFFF if n else 0
            return ".".join(str((mask_int >> (8 * i)) & 0xFF) for i in (3, 2, 1, 0))
    return s


def load_workbook_from_upload(file_storage):
    return openpyxl.load_workbook(io.BytesIO(file_storage.read()), data_only=True)


def header_map(ws, header_row):
    headers = {}
    for idx, cell in enumerate(ws[header_row]):
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = idx
    return headers


def safe_str(v):
    return "" if v is None else str(v).strip()


def build_lookups(wb, opts):
    warnings = []
    if opts["csr_sheet"] not in wb.sheetnames:
        raise ValueError("Sheet not found: " + str(opts["csr_sheet"]) + ". Available: " + ", ".join(wb.sheetnames))
    if opts["ptp_sheet"] not in wb.sheetnames:
        raise ValueError("Sheet not found: " + str(opts["ptp_sheet"]) + ". Available: " + ", ".join(wb.sheetnames))
    header_row = int(opts["header_row"])

    csr_ws = wb[opts["csr_sheet"]]
    csr_headers = header_map(csr_ws, header_row)
    required = [opts["csr_id_col"], opts["esn_col"], opts["lo0_ip_col"],
                opts["lo0_cidr_col"], opts["lo1_ip_col"], opts["lo1_cidr_col"]]
    for col in required:
        if col not in csr_headers:
            raise ValueError("Column not found: '" + str(col) + "' in '" + str(opts["csr_sheet"]) +
                             "' header row " + str(header_row) +
                             ". Available: " + ", ".join(csr_headers))
    idx_csr_id = csr_headers[opts["csr_id_col"]]
    idx_esn = csr_headers[opts["esn_col"]]
    idx_lo0_ip = csr_headers[opts["lo0_ip_col"]]
    idx_lo0_cidr = csr_headers[opts["lo0_cidr_col"]]
    idx_lo1_ip = csr_headers[opts["lo1_ip_col"]]
    idx_lo1_cidr = csr_headers[opts["lo1_cidr_col"]]
    max_idx = max(idx_csr_id, idx_esn, idx_lo0_ip, idx_lo0_cidr, idx_lo1_ip, idx_lo1_cidr)

    esn_to_record = {}
    for row in csr_ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= max_idx:
            continue
        esn = row[idx_esn]
        csr_id = row[idx_csr_id]
        if not (esn and csr_id):
            continue
        key = str(esn).strip()
        record = {
            "csr_id": str(csr_id).strip(),
            "lo0_ip": safe_str(row[idx_lo0_ip]),
            "lo0_cidr": safe_str(row[idx_lo0_cidr]),
            "lo1_ip": safe_str(row[idx_lo1_ip]),
            "lo1_cidr": safe_str(row[idx_lo1_cidr]),
        }
        if key in esn_to_record and esn_to_record[key]["csr_id"] != record["csr_id"]:
            warnings.append("Duplicate ESN " + key)
        esn_to_record[key] = record

    ptp_ws = wb[opts["ptp_sheet"]]
    ptp_headers = header_map(ptp_ws, header_row)
    for col in (opts["ptp_csr_id_col"], opts["ptp_ip_col"], opts["ptp_cidr_col"]):
        if col not in ptp_headers:
            raise ValueError("Column not found: '" + str(col) + "' in '" + str(opts["ptp_sheet"]) +
                             "' header row " + str(header_row) +
                             ". Available: " + ", ".join(ptp_headers))
    p_id_idx = ptp_headers[opts["ptp_csr_id_col"]]
    p_ip_idx = ptp_headers[opts["ptp_ip_col"]]
    p_cidr_idx = ptp_headers[opts["ptp_cidr_col"]]
    p_max = max(p_id_idx, p_ip_idx, p_cidr_idx)

    csrid_to_ptp = {}
    for row in ptp_ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= p_max:
            continue
        csr_id = row[p_id_idx]
        if not csr_id:
            continue
        key = str(csr_id).strip()
        if key not in csrid_to_ptp:
            csrid_to_ptp[key] = (row[p_ip_idx], row[p_cidr_idx])

    return esn_to_record, csrid_to_ptp, warnings


def render_template_text(template, mapping):
    def repl(m):
        key = m.group(1).strip()
        return str(mapping.get(key, m.group(0)))
    return re.sub(r"\{\{\s*(\w+)\s*\}\}", repl, template)


def parse_esn_list(raw):
    if not raw:
        return []
    parts = re.split(r"[\s,;]+", raw.strip())
    return [p for p in (x.strip() for x in parts) if p]


def parse_options(form):
    opts = dict(DEFAULTS)
    for k in opts:
        v = form.get(k)
        if v:
            opts[k] = v
    return opts


def build_rows(esns, esn_to_record, csrid_to_ptp):
    rows = []
    for esn in esns:
        rec = esn_to_record.get(esn)
        if not rec:
            rows.append({
                "esn": esn, "csr_id": None,
                "csr_ip": None, "csr_cidr": None, "mask": None,
                "lo0_ip": None, "lo0_cidr": None, "lo0_mask": None,
                "lo1_ip": None, "lo1_cidr": None, "lo1_mask": None,
                "net": None,
                "status": "ESN not found in CSR List", "ok": False,
            })
            continue
        csr_id = rec["csr_id"]
        ptp = csrid_to_ptp.get(csr_id)
        if not ptp or ptp[0] in (None, ""):
            rows.append({
                "esn": esn, "csr_id": csr_id,
                "csr_ip": None, "csr_cidr": None, "mask": None,
                "lo0_ip": rec["lo0_ip"], "lo0_cidr": rec["lo0_cidr"], "lo0_mask": cidr_to_mask(rec["lo0_cidr"]),
                "lo1_ip": rec["lo1_ip"], "lo1_cidr": rec["lo1_cidr"], "lo1_mask": cidr_to_mask(rec["lo1_cidr"]),
                "net": derive_net_from_ip(rec["lo0_ip"]),
                "status": "No PTP entry for CSR_ID " + csr_id, "ok": False,
            })
            continue
        csr_ip, csr_cidr = ptp
        missing = []
        if not rec["lo0_ip"]:
            missing.append("Lo0_IP")
        if not rec["lo1_ip"]:
            missing.append("Lo1_IP")
        if not csr_ip:
            missing.append("CSR_IP")
        ok = len(missing) == 0
        rows.append({
            "esn": esn,
            "csr_id": csr_id,
            "csr_ip": safe_str(csr_ip),
            "csr_cidr": safe_str(csr_cidr),
            "mask": cidr_to_mask(csr_cidr),
            "lo0_ip": rec["lo0_ip"],
            "lo0_cidr": rec["lo0_cidr"],
            "lo0_mask": cidr_to_mask(rec["lo0_cidr"]),
            "lo1_ip": rec["lo1_ip"],
            "lo1_cidr": rec["lo1_cidr"],
            "lo1_mask": cidr_to_mask(rec["lo1_cidr"]),
            "net": derive_net_from_ip(rec["lo0_ip"]),
            "status": "OK" if ok else ("Missing: " + ", ".join(missing)),
            "ok": ok,
        })
    return rows


@app.route("/")
def index():
    return render_template("index.html", default_template=_load_default_template(), defaults=DEFAULTS)


@app.route("/preview", methods=["POST"])
def preview():
    try:
        if "excel" not in request.files or not request.files["excel"].filename:
            return jsonify({"error": "Please upload an Excel file."}), 400
        wb = load_workbook_from_upload(request.files["excel"])
        opts = parse_options(request.form)
        esns = parse_esn_list(request.form.get("esns", ""))
        if not esns:
            return jsonify({"error": "Please paste at least one ESN."}), 400
        esn_to_record, csrid_to_ptp, warnings = build_lookups(wb, opts)
        rows = build_rows(esns, esn_to_record, csrid_to_ptp)
        return jsonify({
            "rows": rows,
            "warnings": warnings,
            "sheets": wb.sheetnames,
            "summary": {
                "total": len(rows),
                "ok": sum(1 for r in rows if r["ok"]),
                "failed": sum(1 for r in rows if not r["ok"]),
            },
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/generate", methods=["POST"])
def generate():
    try:
        if "excel" not in request.files or not request.files["excel"].filename:
            return ("Please upload an Excel file.", 400)
        wb = load_workbook_from_upload(request.files["excel"])
        opts = parse_options(request.form)
        esns = parse_esn_list(request.form.get("esns", ""))
        template = request.form.get("template", "")
        batch = request.form.get("batch_name", "Batch").strip() or "Batch"
        use_crlf = request.form.get("crlf", "1") == "1"

        if not template.strip():
            return ("Template is empty.", 400)
        if not esns:
            return ("Please paste at least one ESN.", 400)

        esn_to_record, csrid_to_ptp, _ = build_lookups(wb, opts)
        rows = build_rows(esns, esn_to_record, csrid_to_ptp)

        ok_rows = [r for r in rows if r["ok"]]
        bad_rows = [r for r in rows if not r["ok"]]

        if not ok_rows:
            msg_lines = ["No scripts generated. None of the " + str(len(rows)) + " ESN(s) you supplied could be resolved.", "", "Details:"]
            for r in bad_rows[:50]:
                msg_lines.append("  " + r["esn"] + "  ->  " + r["status"])
            if len(bad_rows) > 50:
                msg_lines.append("  ... and " + str(len(bad_rows) - 50) + " more")
            msg_lines.append("")
            msg_lines.append("Check that:")
            msg_lines.append("  - You uploaded the correct IP planning workbook")
            msg_lines.append("  - Sheet names in section 4 match your workbook (case-sensitive)")
            msg_lines.append("  - Each ESN exists in the 'ESN' column of the CSR List tab (paste should be exact, no trailing spaces)")
            return ("\n".join(msg_lines), 400)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        safe_batch = re.sub(r"[^A-Za-z0-9._-]+", "-", batch)
        folder = safe_batch + "_" + timestamp

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            summary_lines = ["\t".join(["ESN", "CSR_ID", "CSR_IP", "MASK",
                                        "Lo0_IP", "Lo0_MASK", "Lo1_IP", "Lo1_MASK",
                                        "NET", "STATUS"])]
            for r in rows:
                summary_lines.append("\t".join([
                    r["esn"], r["csr_id"] or "",
                    r["csr_ip"] or "", r["mask"] or "",
                    r["lo0_ip"] or "", r["lo0_mask"] or "",
                    r["lo1_ip"] or "", r["lo1_mask"] or "",
                    r.get("net") or "",
                    r["status"],
                ]))
                if not r["ok"]:
                    continue
                mapping = {
                    "csr_id": r["csr_id"], "esn": r["esn"],
                    "csr_ip": r["csr_ip"], "csr_cidr": r["csr_cidr"], "mask": r["mask"],
                    "lo0_ip": r["lo0_ip"], "lo0_cidr": r["lo0_cidr"], "lo0_mask": r["lo0_mask"],
                    "lo1_ip": r["lo1_ip"], "lo1_cidr": r["lo1_cidr"], "lo1_mask": r["lo1_mask"],
                    "net": r["net"],
                }
                content = render_template_text(template, mapping)
                if use_crlf:
                    content = content.replace("\r\n", "\n").replace("\n", "\r\n")
                fname = folder + "/" + r["csr_id"] + "_" + r["esn"] + ".txt"
                zf.writestr(fname, content)
            zf.writestr(folder + "/_summary.tsv", "\n".join(summary_lines))

        buf.seek(0)
        resp = send_file(buf, mimetype="application/zip",
                         as_attachment=True, download_name=folder + ".zip")
        resp.headers["X-Generated"] = str(len(ok_rows))
        resp.headers["X-Failed"] = str(len(bad_rows))
        if bad_rows:
            missing = ", ".join((r["esn"] + " (" + r["status"] + ")") for r in bad_rows[:10])
            resp.headers["X-Missing"] = missing[:900]
        resp.headers["Access-Control-Expose-Headers"] = "X-Generated, X-Failed, X-Missing"
        return resp
    except Exception as e:
        return ("Error: " + str(e), 400)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
