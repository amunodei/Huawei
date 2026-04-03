"""
scan_free_ips.py
----------------
Scans management loopback segments for free IP addresses and produces
an Excel report.

Usage (run from a machine inside the network):
    python scan_free_ips.py

Requirements:
    pip install openpyxl

Output:
    free_ips_report.xlsx  (saved in the same folder as this script)
"""

import ipaddress
import subprocess
import platform
import concurrent.futures
from datetime import datetime

# ── Configuration ──────────────────────────────────────────────────────────────
SUBNETS = [
    "10.1.24.0/24",
    "10.1.26.0/24",
]
MAX_WORKERS   = 100   # parallel ping threads
PING_TIMEOUT  = 1     # seconds per ping
OUTPUT_FILE   = "free_ips_report.xlsx"
# ───────────────────────────────────────────────────────────────────────────────


def ping(ip: str) -> bool:
    """Return True if the host responds to a single ICMP ping."""
    system = platform.system().lower()
    if system == "windows":
        cmd = ["ping", "-n", "1", "-w", str(PING_TIMEOUT * 1000), str(ip)]
    else:
        cmd = ["ping", "-c", "1", "-W", str(PING_TIMEOUT), str(ip)]
    try:
        result = subprocess.run(cmd, stdout=subprocess.DEVNULL,
                                stderr=subprocess.DEVNULL, timeout=PING_TIMEOUT + 1)
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        return False


def scan_subnet(subnet_str: str) -> list[dict]:
    """Ping all host addresses in a subnet and return status per IP."""
    network = ipaddress.ip_network(subnet_str, strict=False)
    hosts   = list(network.hosts())
    results = []

    print(f"  Scanning {subnet_str} ({len(hosts)} hosts) …")

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(ping, str(h)): str(h) for h in hosts}
        for future in concurrent.futures.as_completed(futures):
            ip      = futures[future]
            is_up   = future.result()
            results.append({
                "subnet":  subnet_str,
                "ip":      ip,
                "status":  "In Use" if is_up else "Free",
                "last_octets": int(ip.split(".")[-1]),
            })

    results.sort(key=lambda r: r["last_octets"])
    return results


def build_excel(all_results: list[dict], subnets: list[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Palette ────────────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark navy
    FREE_FILL     = PatternFill("solid", start_color="E2EFDA")   # soft green
    INUSE_FILL    = PatternFill("solid", start_color="FCE4D6")   # soft red
    ALT_ROW_FILL  = PatternFill("solid", start_color="F5F5F5")   # light grey
    SUMMARY_FILL  = PatternFill("solid", start_color="D9E1F2")   # light blue

    HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE_FONT    = Font(name="Arial", bold=True, color="1F4E79", size=13)
    BODY_FONT     = Font(name="Arial", size=9)
    BOLD_FONT     = Font(name="Arial", bold=True, size=9)

    thin_side  = Side(style="thin",   color="BFBFBF")
    thick_side = Side(style="medium", color="1F4E79")
    THIN_BORDER  = Border(left=thin_side,  right=thin_side,
                           top=thin_side,  bottom=thin_side)
    THICK_BORDER = Border(left=thick_side, right=thick_side,
                           top=thick_side, bottom=thick_side)

    CENTER  = Alignment(horizontal="center", vertical="center")
    LEFT    = Alignment(horizontal="left",   vertical="center")

    scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Summary Sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 16

    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = "Management Loopback — Free IP Report"
    ws_sum["A1"].font      = TITLE_FONT
    ws_sum["A1"].alignment = CENTER

    ws_sum.merge_cells("A2:D2")
    ws_sum["A2"] = f"Generated: {scan_time}"
    ws_sum["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws_sum["A2"].alignment = CENTER

    ws_sum.row_dimensions[3].height = 6   # spacer

    headers = ["Subnet", "Total Hosts", "Free", "In Use"]
    for col, h in enumerate(headers, start=1):
        cell = ws_sum.cell(row=4, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    row = 5
    for subnet in subnets:
        subnet_rows = [r for r in all_results if r["subnet"] == subnet]
        total  = len(subnet_rows)
        free   = sum(1 for r in subnet_rows if r["status"] == "Free")
        in_use = total - free

        vals = [subnet, total, free, in_use]
        for col, v in enumerate(vals, start=1):
            cell = ws_sum.cell(row=row, column=col, value=v)
            cell.font      = BODY_FONT
            cell.fill      = SUMMARY_FILL
            cell.alignment = CENTER if col > 1 else LEFT
            cell.border    = THIN_BORDER
        row += 1

    # Totals row
    total_all  = len(all_results)
    free_all   = sum(1 for r in all_results if r["status"] == "Free")
    inuse_all  = total_all - free_all
    for col, v in enumerate(["TOTAL", total_all, free_all, inuse_all], start=1):
        cell = ws_sum.cell(row=row, column=col, value=v)
        cell.font      = BOLD_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER if col > 1 else LEFT
        cell.border    = THIN_BORDER
        if col > 1:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        else:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)

    # ── Per-Subnet Sheets ──────────────────────────────────────────────────────
    for subnet in subnets:
        safe_name = subnet.replace("/", "_").replace(".", "-")
        ws = wb.create_sheet(title=safe_name)
        ws.sheet_view.showGridLines = False

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 14

        # Title
        ws.merge_cells("A1:C1")
        ws["A1"] = f"Subnet: {subnet}"
        ws["A1"].font      = TITLE_FONT
        ws["A1"].alignment = CENTER

        ws.merge_cells("A2:C2")
        ws["A2"] = f"Scanned: {scan_time}"
        ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
        ws["A2"].alignment = CENTER

        ws.row_dimensions[3].height = 4

        col_headers = ["IP Address", "Subnet", "Status"]
        for col, h in enumerate(col_headers, start=1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

        subnet_rows = [r for r in all_results if r["subnet"] == subnet]
        for idx, rec in enumerate(subnet_rows):
            data_row = idx + 5
            fill = FREE_FILL if rec["status"] == "Free" else INUSE_FILL
            if rec["status"] != "Free" and idx % 2 == 0:
                fill = INUSE_FILL
            elif rec["status"] == "Free" and idx % 2 != 0:
                fill = FREE_FILL

            for col, val in enumerate([rec["ip"], rec["subnet"], rec["status"]], start=1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.font      = BODY_FONT
                cell.alignment = CENTER
                cell.border    = THIN_BORDER
                if rec["status"] == "Free":
                    cell.fill = FREE_FILL
                else:
                    base = INUSE_FILL if idx % 2 == 0 else PatternFill("solid", start_color="FFF0EB")
                    cell.fill = base

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:C{4 + len(subnet_rows)}"

    # ── Free IPs Only Sheet ────────────────────────────────────────────────────
    ws_free = wb.create_sheet(title="Free IPs Only")
    ws_free.sheet_view.showGridLines = False

    ws_free.column_dimensions["A"].width = 18
    ws_free.column_dimensions["B"].width = 20

    ws_free.merge_cells("A1:B1")
    ws_free["A1"] = "Available IPs for Allocation"
    ws_free["A1"].font      = TITLE_FONT
    ws_free["A1"].alignment = CENTER

    ws_free.merge_cells("A2:B2")
    ws_free["A2"] = f"Scanned: {scan_time}"
    ws_free["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws_free["A2"].alignment = CENTER

    ws_free.row_dimensions[3].height = 4

    for col, h in enumerate(["IP Address", "Subnet"], start=1):
        cell = ws_free.cell(row=4, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    free_rows = [r for r in all_results if r["status"] == "Free"]
    free_rows.sort(key=lambda r: (r["subnet"], r["last_octets"]))

    for idx, rec in enumerate(free_rows):
        data_row = idx + 5
        for col, val in enumerate([rec["ip"], rec["subnet"]], start=1):
            cell = ws_free.cell(row=data_row, column=col, value=val)
            cell.font      = BODY_FONT
            cell.fill      = FREE_FILL if idx % 2 == 0 else PatternFill("solid", start_color="F0FAF0")
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

    ws_free.freeze_panes = "A5"
    ws_free.auto_filter.ref = f"A4:B{4 + len(free_rows)}"

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Report saved: {OUTPUT_FILE}")


def main():
    print("=" * 55)
    print("  Management Loopback — Free IP Scanner")
    print(f"  Subnets : {', '.join(SUBNETS)}")
    print(f"  Workers : {MAX_WORKERS} parallel threads")
    print("=" * 55)

    all_results = []
    for subnet in SUBNETS:
        all_results.extend(scan_subnet(subnet))

    free_count  = sum(1 for r in all_results if r["status"] == "Free")
    inuse_count = len(all_results) - free_count

    print(f"\nResults: {len(all_results)} IPs scanned — "
          f"{free_count} FREE, {inuse_count} in use\n")

    build_excel(all_results, SUBNETS)


if __name__ == "__main__":
    main()