"""
CSR Initial Script Generator - Flask web app.

Takes:
  - A script template (with {{csr_id}}, {{csr_ip}}, {{mask}}, {{csr_cidr}}, {{esn}} placeholders)
  - A list of ESNs
  - An Excel IP planning workbook (with National CSR List + National PTP tabs)
  - A batch name

Produces a zip of one script per ESN, named {csr_id}_{esn}.txt.
"""
from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime
from typing import Any

import openpyxl
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# CIDR -> dotted netmask (covers usual network-engineering cases)
CIDR_MASK = {
    "/30": "255.255.255.252",
    "/31": "255.255.255.254",
    "/32": "255.255.255.255",
    "/29": "255.255.255.248",
    "/28": "255.255.255.240",
    "/27": "255.255.255.224",
    "/26": "255.255.255.192",
    "/25": "255.255.255.128",
    "/24": "255.255.255.0",
}


DEFAULT_TEMPLATE = """display esn

disp license resource usage

sysname {{csr_id}}

  #
  license
active port-basic slot 2 port 10-11
  #
  interface GigabitEthernet0/2/10
 port-mode 10GE
 Y
  #
  interface GigabitEthernet0/2/11
 port-mode 10GE
 Y
 #
 #
isis 10
 is-level level-2
 #
mpls
#
mpls ldp
 #
 interface GigabitEthernet0/2/8
 carrier down-hold-time 200
 mtu 9216
 description CSR uplink
 undo shutdown
 set flow-stat interval 10
 trap-threshold input-rate 80 resume-rate 80
 trap-threshold output-rate 80 resume-rate 80
 ip address {{csr_ip}} {{mask}}
 trust upstream default
 isis enable 10
 isis circuit-type p2p
 isis circuit-level level-2
 isis ldp-sync
 mpls
 mpls ldp
 dcn
 statistic enable
#
commit




"""


DEFAULTS = {
    "csr_sheet": "National CSR List",
    "ptp_sheet": "National PTP",
    "csr_id_col": "CSR_ID",
    "esn_col": "ESN",
    "ptp_csr_id_col": "CSR_ID",
    "ptp_ip_col": "CSR_IP",
    "ptp_cidr_col": "CSR_CIDR",
    "header_row": 2,
}


def cidr_to_mask(cidr: Any) -> str:
    if cidr is None:
        return ""
    s = str(cidr).strip()
    if s in CIDR_MASK:
        return CIDR_MASK[s]
    # If cidr is already dotted, pass through
    if "." in s:
        return s
    # Derive from prefix length if given as /N or N
    m = re.match(r"/?(\d{1,2})$", s)
    if m:
        n = int(m.group(1))
        if 0 <= n <= 32:
            mask_int = (0xFFFFFFFF << (32 - n)) & 0xFFFFFFFF if n else 0
            return ".".join(str((mask_int >> (8 * i)) & 0xFF) for i in (3, 2, 1, 0))
    return s


def load_workbook_from_upload(file_storage) -> openpyxl.Workbook:
    data = file_storage.read()
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def header_map(ws, header_row: int) -> dict[str, int]:
    headers = {}
    for idx, cell in enumerate(ws[header_row]):
        val = cell.value
        if val is None:
            continue
        headers[str(val).strip()] = idx
    return headers


def build_lookups(wb, opts: dict) -> tuple[dict, dict, list[str]]:
    """Returns (esn_to_csrid, csrid_to_ip, warnings)."""
    warnings: list[str] = []

    csr_sheet = opts["csr_sheet"]
    ptp_sheet = opts["ptp_sheet"]
    if csr_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{csr_sheet}' not found. Available: {wb.sheetnames}")
    if ptp_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{ptp_sheet}' not found. Available: {wb.sheetnames}")

    header_row = int(opts["header_row"])

    # CSR list
    csr_ws = wb[csr_sheet]
    csr_headers = header_map(csr_ws, header_row)
    for col in (opts["csr_id_col"], opts["esn_col"]):
        if col not in csr_headers:
            raise ValueError(
                f"Column '{col}' not found in '{csr_sheet}' header row {header_row}. "
                f"Available: {list(csr_headers)}"
            )
    csr_id_idx = csr_headers[opts["csr_id_col"]]
    esn_idx = csr_headers[opts["esn_col"]]

    esn_to_csrid: dict[str, str] = {}
    for row in csr_ws.iter_rows(min_row=header_row + 1, values_only=True):
        if esn_idx >= len(row) or csr_id_idx >= len(row):
            continue
        esn = row[esn_idx]
        csr_id = row[csr_id_idx]
        if esn and csr_id:
            key = str(esn).strip()
            if key in esn_to_csrid and esn_to_csrid[key] != csr_id:
                warnings.append(
                    f"Duplicate ESN {key} in '{csr_sheet}': {esn_to_csrid[key]} vs {csr_id}"
                )
            esn_to_csrid[key] = str(csr_id).strip()

    # PTP
    ptp_ws = wb[ptp_sheet]
    ptp_headers = header_map(ptp_ws, header_row)
    for col in (opts["ptp_csr_id_col"], opts["ptp_ip_col"], opts["ptp_cidr_col"]):
        if col not in ptp_headers:
            raise ValueError(
                f"Column '{col}' not found in '{ptp_sheet}' header row {header_row}. "
                f"Available: {list(ptp_headers)}"
            )
    p_id_idx = ptp_headers[opts["ptp_csr_id_col"]]
    p_ip_idx = ptp_headers[opts["ptp_ip_col"]]
    p_cidr_idx = ptp_headers[opts["ptp_cidr_col"]]

    csrid_to_ip: dict[str, tuple[Any, Any]] = {}
    for row in ptp_ws.iter_rows(min_row=header_row + 1, values_only=True):
        if max(p_id_idx, p_ip_idx, p_cidr_idx) >= len(row):
            continue
        csr_id = row[p_id_idx]
        if not csr_id:
            continue
        key = str(csr_id).strip()
        # Keep first occurrence (typical for the "main" uplink row)
        if key not in csrid_to_ip:
            csrid_to_ip[key] = (row[p_ip_idx], row[p_cidr_idx])

    return esn_to_csrid, csrid_to_ip, warnings


def render_template_text(template: str, mapping: dict[str, str]) -> str:
    """Replace {{key}} placeholders. Unknown keys are left as-is."""
    def repl(m: re.Match) -> str:
        key = m.group(1).strip()
        return str(mapping.get(key, m.group(0)))
    return re.sub(r"\{\{\s*(\w+)\s*\}\}", repl, template)


def parse_esn_list(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[\s,;]+", raw.strip())
    return [p for p in (x.strip() for x in parts) if p]


def parse_options(form) -> dict:
    opts = dict(DEFAULTS)
    for k in opts:
        v = form.get(k)
        if v:
            opts[k] = v
    return opts


def build_rows(template: str, esns: list[str], esn_to_csrid: dict, csrid_to_ip: dict):
    rows = []
    for esn in esns:
        csr_id = esn_to_csrid.get(esn)
        if not csr_id:
            rows.append({
                "esn": esn, "csr_id": None, "csr_ip": None, "csr_cidr": None,
                "mask": None, "status": "ESN not found in CSR List", "ok": False,
            })
            continue
        ip_info = csrid_to_ip.get(csr_id)
        if not ip_info or ip_info[0] is None:
            rows.append({
                "esn": esn, "csr_id": csr_id, "csr_ip": None, "csr_cidr": None,
                "mask": None, "status": "No PTP entry for CSR_ID", "ok": False,
            })
            continue
        csr_ip, csr_cidr = ip_info
        mask = cidr_to_mask(csr_cidr)
        rows.append({
            "esn": esn,
            "csr_id": csr_id,
            "csr_ip": str(csr_ip).strip() if csr_ip else "",
            "csr_cidr": str(csr_cidr).strip() if csr_cidr else "",
            "mask": mask,
            "status": "OK",
            "ok": True,
        })
    return rows


@app.route("/")
def index():
    return render_template("index.html", default_template=DEFAULT_TEMPLATE, defaults=DEFAULTS)


@app.route("/preview", methods=["POST"])
def preview():
    try:
        if "excel" not in request.files or not request.files["excel"].filename:
            return jsonify({"error": "Please upload an Excel file."}), 400
        wb = load_workbook_from_upload(request.files["excel"])
        opts = parse_options(request.form)
        esns = parse_esn_list(request.form.get("esns", ""))
        if not esns:
            return jsonify({"error": "Please paste at least one ESN."}), 400

        esn_to_csrid, csrid_to_ip, warnings = build_lookups(wb, opts)
        rows = build_rows("", esns, esn_to_csrid, csrid_to_ip)

        return jsonify({
            "rows": rows,
            "warnings": warnings,
            "sheets": wb.sheetnames,
            "summary": {
                "total": len(rows),
                "ok": sum(1 for r in rows if r["ok"]),
                "failed": sum(1 for r in rows if not r["ok"]),
            },
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/generate", methods=["POST"])
def generate():
    try:
        if "excel" not in request.files or not request.files["excel"].filename:
            return ("Please upload an Excel file.", 400)
        wb = load_workbook_from_upload(request.files["excel"])
        opts = parse_options(request.form)
        esns = parse_esn_list(request.form.get("esns", ""))
        template = request.form.get("template", "")
        batch = request.form.get("batch_name", "Batch").strip() or "Batch"
        use_crlf = request.form.get("crlf", "1") == "1"

        if not template.strip():
            return ("Template is empty.", 400)
        if not esns:
            return ("Please paste at least one ESN.", 400)

        esn_to_csrid, csrid_to_ip, _ = build_lookups(wb, opts)
        rows = build_rows(template, esns, esn_to_csrid, csrid_to_ip)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        folder = f"{re.sub(r'[^A-Za-z0-9._-]+', '-', batch)}_{timestamp}"

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            summary_lines = ["ESN\tCSR_ID\tCSR_IP\tMASK\tSTATUS"]
            for r in rows:
                summary_lines.append(
                    f"{r['esn']}\t{r['csr_id'] or ''}\t{r['csr_ip'] or ''}\t{r['mask'] or ''}\t{r['status']}"
                )
                if not r["ok"]:
                    continue
                mapping = {
                    "csr_id": r["csr_id"],
                    "csr_ip": r["csr_ip"],
                    "csr_cidr": r["csr_cidr"],
                    "mask": r["mask"],
                    "esn": r["esn"],
                }
                content = render_template_text(template, mapping)
                if use_crlf:
                    content = content.replace("\r\n", "\n").replace("\n", "\r\n")
                fname = f"{folder}/{r['csr_id']}_{r['esn']}.txt"
                zf.writestr(fname, content)
            zf.writestr(f"{folder}/_summary.tsv", "\n".join(summary_lines))

        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"{folder}.zip",
        )
    except Exception as e:
        return (f"Error: {e}", 400)


if __name__ == "__main__":
    # Bind to localhost; change to 0.0.0.0 if you want LAN access
    app.run(host="127.0.0.1", port=5000, debug=False)
